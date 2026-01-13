from flask import Flask, render_template, jsonify, request, send_file
import json
import os
import re
import openpyxl
from io import BytesIO
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Constants
CONTACTS_FILE = "data.json"
DEPARTMENTS_FILE = "departments.json"
ASSIGNMENTS_FILE = "assignments.json"
CV_FOLDER = "cv_files"

# Ensure required directories and files exist
if not os.path.exists(CV_FOLDER):
    os.makedirs(CV_FOLDER)


# ==================== DATA MANAGEMENT FUNCTIONS ====================

def load_data():
    """Load contacts data from JSON file"""
    if os.path.exists(CONTACTS_FILE):
        try:
            return sorted(json.load(open(CONTACTS_FILE, "r", encoding="utf-8")),
                          key=lambda x: x["name"].lower())
        except json.JSONDecodeError:
            return []
    return []


def save_data(data):
    """Save contacts data to JSON file"""
    data = sorted(data, key=lambda x: x["name"].lower())
    with open(CONTACTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def load_departments():
    """Load departments data from JSON file"""
    if os.path.exists(DEPARTMENTS_FILE):
        try:
            return sorted(json.load(open(DEPARTMENTS_FILE, "r", encoding="utf-8")),
                          key=lambda x: x["name"].lower())
        except json.JSONDecodeError:
            return []
    return []


def save_departments(data):
    """Save departments data to JSON file"""
    data = sorted(data, key=lambda x: x["name"].lower())
    with open(DEPARTMENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def load_assignments():
    """Load assignments data from JSON file"""
    if os.path.exists(ASSIGNMENTS_FILE):
        try:
            return json.load(open(ASSIGNMENTS_FILE, "r", encoding="utf-8"))
        except json.JSONDecodeError:
            return []
    return []


def save_assignments(data):
    """Save assignments data to JSON file"""
    with open(ASSIGNMENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


# ==================== ROUTES ====================

@app.route("/")
def dashboard():
    """Main dashboard page"""
    return render_template("dashboard.html")


# ==================== CONTACTS ROUTES ====================

@app.route("/contacts")
def contacts_page():
    """Contacts management page - show ALL contacts"""
    data = load_data()
    return render_template("contacts.html", data=data)


@app.route("/api/data")
def api_data():
    """API endpoint for ALL contacts data"""
    data = load_data()
    return jsonify(data)


@app.route("/add", methods=["POST"])
def add_contact():
    """Add or edit a contact"""
    data = load_data()
    new_contact = request.json
    name = new_contact.get("name", "").strip()
    phone = new_contact.get("phone", "").strip()
    location = new_contact.get("location", "").strip()
    old_name = new_contact.get("old_name", "").strip()
    old_phone = new_contact.get("old_phone", "").strip()
    status = new_contact.get("status", "waiting").strip()

    # Validation
    name_regex = re.compile(r"^[A-Za-zΑ-Ωα-ωΆΈΊΌΎΏΉάέίόύώή ]{1,30}$")
    phone_regex = re.compile(r"^69[0-9]{8}$")

    if not name_regex.match(name):
        return jsonify({"error": "Invalid name"}), 400
    if not phone_regex.match(phone):
        return jsonify({"error": "Invalid phone"}), 400

    # Check for duplicate phone number
    for c in data:
        if c["phone"] == phone:
            if old_phone and c["phone"] == old_phone:
                continue
            return jsonify({"error": "Phone number exists"}), 400

    # Edit existing contact
    if old_name and old_phone:
        for i, c in enumerate(data):
            if c["name"].lower() == old_name.lower() and c["phone"] == old_phone:
                data[i] = {"name": name, "phone": phone, "location": location, "status": status}
                save_data(data)
                return jsonify({"success": True})

    # Add new contact
    data.append({"name": name, "phone": phone, "location": location, "status": status})
    save_data(data)
    return jsonify({"success": True})


@app.route("/delete", methods=["POST"])
def delete_contacts():
    """Delete contacts and their associated data"""
    data = load_data()
    phones_to_delete = request.json.get("phones", [])

    # Delete associated CV files
    for contact in data:
        if contact["phone"] in phones_to_delete:
            full_name = contact["name"]
            safe_name = re.sub(r'[^A-Za-zΑ-Ωα-ωΆΈΊΌΎΏΉάέίόύώή0-9\s\-_]', '', full_name)
            safe_name = safe_name.replace(' ', '_')

            for file in os.listdir(CV_FOLDER):
                if file.startswith(f"CV_{safe_name}_") and file.endswith(".pdf"):
                    try:
                        os.remove(os.path.join(CV_FOLDER, file))
                        print(f"Deleted CV: {file}")
                    except Exception as e:
                        print(f"Error deleting CV file {file}: {e}")

    # Update assignments
    assignments = load_assignments()
    contacts_to_keep = [c for c in data if c["phone"] not in phones_to_delete]
    phone_to_index = {c["phone"]: i for i, c in enumerate(contacts_to_keep)}
    assignments = [a for a in assignments if a.get("contact_index") in phone_to_index.values()]
    save_assignments(assignments)

    # Delete the contacts
    data = contacts_to_keep
    save_data(data)

    return jsonify({"success": True})


# ==================== ACTIVE CONTACTS ROUTES ====================

@app.route("/active_contacts")
def active_contacts_page():
    """Active contacts management page"""
    departments = load_departments()
    return render_template("active_contacts.html", departments=departments)


@app.route("/api/active_contacts")
def api_active_contacts():
    """API endpoint for active contacts data"""
    contacts = load_data()
    departments = load_departments()
    assignments = load_assignments()

    # Group active contacts by department
    active_contacts_by_dept = {}

    for dept_idx, department in enumerate(departments):
        dept_contacts = []

        # Find active contacts assigned to this department
        for assignment in assignments:
            if assignment["dept_index"] == dept_idx:
                contact_idx = assignment["contact_index"]
                if contact_idx < len(contacts) and contacts[contact_idx].get("status") == "active":
                    contact = contacts[contact_idx]
                    dept_contacts.append({
                        "name": contact["name"],
                        "phone": contact["phone"],
                        "location": contact.get("location", "")
                    })

        if dept_contacts:  # Only include departments with active contacts
            active_contacts_by_dept[department["name"]] = dept_contacts

    return jsonify(active_contacts_by_dept)


@app.route("/move_to_active", methods=["POST"])
def move_to_active():
    """Move contact to active status"""
    data = request.json
    phone = data.get("phone")

    contacts = load_data()

    # Find and update the contact
    for contact in contacts:
        if contact["phone"] == phone:
            contact["status"] = "active"
            break

    save_data(contacts)
    return jsonify({"success": True})


@app.route("/move_to_waiting", methods=["POST"])
def move_to_waiting():
    """Move contact to waiting status"""
    data = request.json
    phone = data.get("phone")

    contacts = load_data()

    # Find and update the contact
    for contact in contacts:
        if contact["phone"] == phone:
            contact["status"] = "waiting"
            break

    save_data(contacts)
    return jsonify({"success": True})


@app.route("/move_to_inactive", methods=["POST"])
def move_to_inactive():
    """Move contact to inactive status"""
    data = request.json
    phone = data.get("phone")

    contacts = load_data()

    # Find and update the contact
    for contact in contacts:
        if contact["phone"] == phone:
            contact["status"] = "inactive"
            break

    save_data(contacts)
    return jsonify({"success": True})


# ==================== DEPARTMENTS ROUTES ====================

@app.route("/departments")
def departments_page():
    """Departments management page"""
    data = load_departments()
    return render_template("departments.html", data=data)


@app.route("/api/departments")
def api_departments():
    """API endpoint for departments data"""
    return jsonify(load_departments())


@app.route("/departments/add", methods=["POST"])
def add_department():
    """Add or edit a department"""
    data = load_departments()
    new_dept = request.json
    name = new_dept.get("name", "").strip()
    old_name = new_dept.get("old_name", "").strip()

    if not name:
        return jsonify({"error": "Invalid name"}), 400

    # Edit existing department
    if old_name:
        for i, d in enumerate(data):
            if d["name"].lower() == old_name.lower():
                data[i] = {"name": name}
                save_departments(data)
                return jsonify({"success": True})

    # Add new department
    data.append({"name": name})
    save_departments(data)
    return jsonify({"success": True})


@app.route("/departments/delete", methods=["POST"])
def delete_department():
    """Delete departments and update assignments"""
    data = load_departments()
    names_to_delete = request.json.get("names", [])

    # Find indices of departments to delete
    indices_to_delete = [i for i, d in enumerate(data) if d["name"] in names_to_delete]

    # Update assignments
    assignments = load_assignments()
    assignments = [a for a in assignments if a.get("dept_index") not in indices_to_delete]

    # Reindex remaining assignments for departments that come after deleted ones
    for assignment in assignments:
        dept_idx = assignment["dept_index"]
        # Count how many deleted departments come before this department index
        shift_amount = sum(1 for idx in indices_to_delete if idx < dept_idx)
        assignment["dept_index"] -= shift_amount

    save_assignments(assignments)

    # Delete the departments
    data = [d for d in data if d["name"] not in names_to_delete]
    save_departments(data)

    return jsonify({"success": True})


# ==================== ASSIGNMENTS ROUTES ====================

@app.route("/assignments")
def assignments_page():
    """Assignments management page"""
    contacts = load_data()
    departments = load_departments()
    return render_template("assignments.html", contacts=contacts, departments=departments)


@app.route("/api/assignments")
def api_assignments():
    """API endpoint for assignments data"""
    assignments = load_assignments()
    contacts = load_data()
    departments = load_departments()
    result = []

    for a in assignments:
        c_idx = a.get("contact_index")
        d_idx = a.get("dept_index")
        if c_idx is None or d_idx is None:
            continue
        result.append({
            "contact_name": contacts[c_idx]["name"] if c_idx < len(contacts) else "Unknown",
            "department_name": departments[d_idx]["name"] if d_idx < len(departments) else "Unknown"
        })

    # Sort by department name alphabetically
    result.sort(key=lambda x: x["department_name"].lower())

    return jsonify(result)


@app.route("/assignments/add", methods=["POST"])
def add_assignment():
    """Add a new assignment - supports both index-based and name-based"""
    data = request.json
    contact_index = data.get("contact_index")
    dept_index = data.get("dept_index")
    contact_phone = data.get("contact_phone")
    department_name = data.get("department_name")

    contacts = load_data()
    departments = load_departments()
    assignments = load_assignments()

    # Handle name-based assignment (new method)
    if contact_phone and department_name:
        # Find contact index by phone
        contact_index = None
        for i, contact in enumerate(contacts):
            if contact["phone"] == contact_phone:
                contact_index = i
                break

        # Find department index by name
        dept_index = None
        for i, department in enumerate(departments):
            if department["name"] == department_name:
                dept_index = i
                break

        if contact_index is None:
            return jsonify({"error": "Contact not found"}), 400
        if dept_index is None:
            return jsonify({"error": "Department not found"}), 400

    # Handle index-based assignment (old method)
    elif contact_index is not None and dept_index is not None:
        # Validate indices
        if contact_index >= len(contacts) or dept_index >= len(departments):
            return jsonify({"error": "Invalid contact or department index"}), 400
    else:
        return jsonify({"error": "Missing selection"}), 400

    # Prevent duplicates
    for a in assignments:
        if a["contact_index"] == contact_index and a["dept_index"] == dept_index:
            return jsonify({"error": "This assignment already exists"}), 400

    assignments.append({"contact_index": contact_index, "dept_index": dept_index})
    save_assignments(assignments)
    return jsonify({"success": True})


@app.route("/assignments/delete", methods=["POST"])
def delete_assignment():
    """Delete an assignment"""
    data = request.json
    index = data.get("index")

    if index is None:
        return jsonify({"error": "Missing index"}), 400

    assignments = load_assignments()
    if 0 <= index < len(assignments):
        assignments.pop(index)
        save_assignments(assignments)

    return jsonify({"success": True})


@app.route("/assignments/add_by_name", methods=["POST"])
def add_assignment_by_name():
    """Add assignment using contact phone and department name"""
    data = request.json
    contact_phone = data.get("contact_phone")
    department_name = data.get("department_name")

    if not contact_phone or not department_name:
        return jsonify({"error": "Missing contact or department"}), 400

    contacts = load_data()
    departments = load_departments()
    assignments = load_assignments()

    # Find contact index by phone
    contact_index = None
    for i, contact in enumerate(contacts):
        if contact["phone"] == contact_phone:
            contact_index = i
            break

    # Find department index by name
    dept_index = None
    for i, department in enumerate(departments):
        if department["name"] == department_name:
            dept_index = i
            break

    if contact_index is None:
        return jsonify({"error": "Contact not found"}), 400
    if dept_index is None:
        return jsonify({"error": "Department not found"}), 400

    # Prevent duplicates
    for a in assignments:
        if a["contact_index"] == contact_index and a["dept_index"] == dept_index:
            return jsonify({"error": "This assignment already exists"}), 400

    assignments.append({"contact_index": contact_index, "dept_index": dept_index})
    save_assignments(assignments)
    return jsonify({"success": True})


@app.route("/assignments/delete_by_name", methods=["POST"])
def delete_assignment_by_name():
    """Delete assignment using contact name and department name"""
    data = request.json
    contact_name = data.get("contact_name")
    department_name = data.get("department_name")

    if not contact_name or not department_name:
        return jsonify({"error": "Missing contact or department"}), 400

    contacts = load_data()
    departments = load_departments()
    assignments = load_assignments()

    # Find the assignment to delete
    assignment_to_delete = None
    for i, assignment in enumerate(assignments):
        contact_idx = assignment.get("contact_index")
        dept_idx = assignment.get("dept_index")

        if (contact_idx is not None and dept_idx is not None and
                contact_idx < len(contacts) and dept_idx < len(departments) and
                contacts[contact_idx]["name"] == contact_name and
                departments[dept_idx]["name"] == department_name):
            assignment_to_delete = i
            break

    if assignment_to_delete is not None:
        assignments.pop(assignment_to_delete)
        save_assignments(assignments)
        return jsonify({"success": True})
    else:
        return jsonify({"error": "Assignment not found"}), 404
# ==================== EXPORT/IMPORT ROUTES ====================

@app.route("/export")
def export_excel():
    """Export all data to Excel file"""
    contacts = load_data()
    departments = load_departments()
    assignments = load_assignments()

    # Create lookup dictionary for assignments
    contact_to_departments = {}  # contact_index -> list of department names
    for a in assignments:
        c_idx = a.get("contact_index")
        d_idx = a.get("dept_index")
        if c_idx is not None and d_idx is not None:
            dept_name = departments[d_idx]["name"] if d_idx < len(departments) else ""
            contact_to_departments.setdefault(c_idx, []).append(dept_name)

    wb = openpyxl.Workbook()

    # ---------------- Contacts sheet ----------------
    ws1 = wb.active
    ws1.title = "Contacts"
    ws1.append(["Name", "Phone", "Status", "Location", "Departments"])

    for idx, c in enumerate(contacts):
        dept_list = contact_to_departments.get(idx, [])
        dept_str = ", ".join(dept_list)
        ws1.append([c.get("name", ""), c.get("phone", ""), c.get("status", "waiting"), c.get("location", ""), dept_str])

    # ---------------- Departments sheet ----------------
    ws2 = wb.create_sheet("Departments")
    ws2.append(["Name", "Contact Count"])

    # Calculate how many contacts are assigned to each department
    dept_contact_count = {}
    for a in assignments:
        d_idx = a.get("dept_index")
        if d_idx is not None and d_idx < len(departments):
            dept_contact_count[d_idx] = dept_contact_count.get(d_idx, 0) + 1

    for idx, d in enumerate(departments):
        contact_count = dept_contact_count.get(idx, 0)
        ws2.append([d.get("name", ""), contact_count])

    # ---------------- Assignments sheet ----------------
    ws3 = wb.create_sheet("Assignments")
    ws3.append(["Contact", "Department", "Contact Status"])
    for a in assignments:
        c_idx = a.get("contact_index")
        d_idx = a.get("dept_index")
        contact_name = contacts[c_idx]["name"] if c_idx is not None and 0 <= c_idx < len(contacts) else ""
        dept_name = departments[d_idx]["name"] if d_idx is not None and 0 <= d_idx < len(departments) else ""
        contact_status = contacts[c_idx].get("status", "waiting") if c_idx is not None and 0 <= c_idx < len(
            contacts) else ""
        ws3.append([contact_name, dept_name, contact_status])

    # ---------------- Summary sheet ----------------
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Total Contacts", len(contacts)])
    ws4.append(["Total Departments", len(departments)])
    ws4.append(["Total Assignments", len(assignments)])

    # Status breakdown
    status_count = {}
    for contact in contacts:
        status = contact.get("status", "waiting")
        status_count[status] = status_count.get(status, 0) + 1

    ws4.append([])
    ws4.append(["Status Breakdown"])
    for status, count in status_count.items():
        ws4.append([status.title(), count])

    # ---------------- Save to BytesIO ----------------
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/import", methods=["POST"])
def import_excel():
    """Import data from Excel file"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        if not file.filename.endswith(".xlsx"):
            return jsonify({"error": "Invalid file type"}), 400

        wb = openpyxl.load_workbook(file)
        skipped = []

        # ---------------- Departments ----------------
        departments = load_departments()
        existing_dept_names = {d["name"] for d in departments}

        if "Departments" in wb.sheetnames:
            ws = wb["Departments"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                name = str(row[0] or "").strip()
                if name and name not in existing_dept_names:
                    departments.append({"name": name})
                    existing_dept_names.add(name)

            save_departments(departments)

        dept_name_to_idx = {d["name"]: idx for idx, d in enumerate(departments)}

        # ---------------- Contacts + Assignments ----------------
        existing_contacts = load_data()
        assignments = load_assignments()

        # Create phone to index mapping for existing contacts
        phone_to_index = {c["phone"]: idx for idx, c in enumerate(existing_contacts)}

        name_regex = re.compile(r"^[A-Za-zΑ-Ωα-ωΆΈΊΌΎΏΉάέίόύώή ]{1,30}$")
        phone_regex = re.compile(r"^69[0-9]{8}$")

        if "Contacts" in wb.sheetnames:
            ws = wb["Contacts"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                try:
                    name = str(row[0] or "").strip()
                    phone = str(row[1] or "").strip()
                    status = str(row[2] or "").strip().lower() if len(row) > 2 and row[2] else "waiting"
                    location = str(row[3] or "").strip() if len(row) > 3 and row[3] else ""
                    dept_str = str(row[4] or "").strip() if len(row) > 4 and row[4] else ""

                    # Validation - name and phone are required
                    if not (name and phone):
                        skipped.append(f"Row {row_idx + 2}: {name} ({phone}) - Missing name or phone")
                        continue
                    if not name_regex.match(name):
                        skipped.append(f"Row {row_idx + 2}: {name} ({phone}) - Invalid name")
                        continue
                    if not phone_regex.match(phone):
                        skipped.append(f"Row {row_idx + 2}: {name} ({phone}) - Invalid phone")
                        continue

                    # Validate status
                    if status not in ["active", "inactive", "waiting"]:
                        status = "waiting"

                    # Check if contact exists
                    if phone in phone_to_index:
                        # Update existing contact
                        contact_idx = phone_to_index[phone]
                        existing_contacts[contact_idx].update({
                            "name": name,
                            "status": status,
                            "location": location
                        })
                    else:
                        # Add new contact
                        existing_contacts.append({
                            "name": name,
                            "phone": phone,
                            "status": status,
                            "location": location
                        })
                        contact_idx = len(existing_contacts) - 1
                        phone_to_index[phone] = contact_idx

                    # Process departments
                    if dept_str:
                        dept_names = [d.strip() for d in dept_str.split(",") if d.strip()]
                        for dept_name in dept_names:
                            if dept_name in dept_name_to_idx:
                                dept_idx = dept_name_to_idx[dept_name]
                                # Check if assignment already exists
                                if not any(a["contact_index"] == contact_idx and a["dept_index"] == dept_idx for a in
                                           assignments):
                                    assignments.append({
                                        "contact_index": contact_idx,
                                        "dept_index": dept_idx
                                    })
                            else:
                                skipped.append(f"Row {row_idx + 2}: {name} ({phone}) - Unknown department: {dept_name}")

                except Exception as row_err:
                    skipped.append(f"Row {row_idx + 2} error: {str(row_err)}")

        # Save final data
        save_data(existing_contacts)
        save_assignments(assignments)

        return jsonify({"success": True, "skipped": skipped})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ==================== CV MANAGEMENT ROUTES ====================

@app.route("/contacts/upload_cv", methods=["POST"])
def upload_cv():
    """Upload CV file for a contact"""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".pdf"):
        return jsonify({"error": "Invalid file type"}), 400

    full_name = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()

    # Keep Greek characters and basic symbols
    safe_name = re.sub(r'[^A-Za-zΑ-Ωα-ωΆΈΊΌΎΏΉάέίόύώή0-9\s\-_]', '', full_name)
    safe_name = safe_name.replace(' ', '_')

    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"CV_{safe_name}_{date_str}.pdf"

    # Delete any existing CV files for this contact
    for existing_file in os.listdir(CV_FOLDER):
        if existing_file.startswith(f"CV_{safe_name}_") and existing_file.endswith(".pdf"):
            try:
                os.remove(os.path.join(CV_FOLDER, existing_file))
            except Exception as e:
                print(f"Error deleting old CV {existing_file}: {e}")

    file.save(os.path.join(CV_FOLDER, filename))
    return jsonify({"success": True, "filename": filename})


@app.route("/cv/<name>")
def view_cv(name):
    """View CV file for a contact"""
    import urllib.parse
    decoded_name = urllib.parse.unquote(name)

    safe_name = re.sub(r'[^A-Za-zΑ-Ωα-ωΆΈΊΌΎΏΉάέίόύώή0-9\s\-_]', '', decoded_name)
    safe_name = safe_name.replace(' ', '_')

    # Find the most recent CV file for this contact
    cv_files = []
    for file in os.listdir(CV_FOLDER):
        if file.startswith(f"CV_{safe_name}_") and file.endswith(".pdf"):
            cv_files.append(file)

    # Sort by date (newest first) and return the most recent
    if cv_files:
        cv_files.sort(reverse=True)
        return send_file(os.path.join(CV_FOLDER, cv_files[0]))

    return "CV not found", 404


# ==================== APPLICATION INITIALIZATION ====================

if __name__ == "__main__":
    # Ensure required files exist
    for file in [CONTACTS_FILE, DEPARTMENTS_FILE, ASSIGNMENTS_FILE]:
        if not os.path.exists(file):
            with open(file, "w", encoding="utf-8") as f:
                f.write("[]")

    # Ensure CV folder exists
    if not os.path.exists(CV_FOLDER):
        os.makedirs(CV_FOLDER)

    # Start the application
    app.run(port=5020, debug=True)